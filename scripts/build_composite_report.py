#!/usr/bin/env python3
"""Generate Excel report from composite_result_*.json files.

Run AFTER:
  node scripts/composite_strategy_lab.mjs
  node scripts/build_composite_presets.mjs

Usage: python scripts/build_composite_report.py
Output: data/coin_lab/composite_strategy_report.xlsx
"""
import json, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "coin_lab"
OUT = DATA / "composite_strategy_report.xlsx"
PERIODS = ["1day", "4hour", "60min"]
DROPPED = {"BCH", "ZEC"}

ARIAL = Font(name="Arial", size=10)
ARIAL_BOLD = Font(name="Arial", size=10, bold=True)
ARIAL_HEAD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", start_color="305496")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
GOOD_FILL = PatternFill("solid", start_color="C6EFCE")
BAD_FILL = PatternFill("solid", start_color="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = ARIAL_HEAD
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def load_period(period):
    f = DATA / f"composite_result_{period}.json"
    if not f.exists():
        # fallback to legacy single file for 1day
        if period == "1day":
            f = DATA / "composite_result.json"
            if not f.exists():
                return None
        else:
            return None
    return json.loads(f.read_text())


def write_summary(ws, all_data):
    ws.title = "Summary"
    headers = ["Period", "Coin", "Symbol", "Candles", "Lows", "Highs",
               "Train Ret", "Train Sharpe", "Test Ret", "Test Sharpe",
               "Full Ret", "Full CAGR", "Full MDD", "Trades", "Win%",
               "Strategies", "Threshold"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    row = 2
    for period in PERIODS:
        data = all_data.get(period)
        if not data:
            continue
        for coin, d in data.items():
            if coin in DROPPED:
                continue
            bt = d["backtest"]
            strats = ",".join(f"{k}({v:.2f})" for k, v in d["weights"].items())
            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=coin)
            ws.cell(row=row, column=3, value=d["symbol"])
            ws.cell(row=row, column=4, value=d["candles"])
            ws.cell(row=row, column=5, value=d["swings"]["lows"])
            ws.cell(row=row, column=6, value=d["swings"]["highs"])
            ws.cell(row=row, column=7, value=bt["train"]["totalReturn"]).number_format = "0.0%"
            ws.cell(row=row, column=8, value=bt["train"]["sharpe"]).number_format = "0.00"
            ws.cell(row=row, column=9, value=bt["test"]["totalReturn"]).number_format = "0.0%"
            ws.cell(row=row, column=10, value=bt["test"]["sharpe"]).number_format = "0.00"
            ws.cell(row=row, column=11, value=bt["full"]["totalReturn"]).number_format = "0.0%"
            ws.cell(row=row, column=12, value=bt["full"]["cagr"]).number_format = "0.0%"
            ws.cell(row=row, column=13, value=bt["full"]["mdd"]).number_format = "0.0%"
            ws.cell(row=row, column=14, value=bt["full"]["trades"])
            ws.cell(row=row, column=15, value=bt["full"]["winRate"]).number_format = "0%"
            ws.cell(row=row, column=16, value=strats)
            ws.cell(row=row, column=17, value=d["threshold"]).number_format = "0.00"
            # color test sharpe
            ts = bt["test"]["sharpe"]
            if ts >= 0.5:
                ws.cell(row=row, column=10).fill = GOOD_FILL
            elif ts < 0:
                ws.cell(row=row, column=10).fill = BAD_FILL
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = ARIAL
                cell.border = BORDER
                if c >= 4 and c <= 15:
                    cell.alignment = CENTER
                if row % 2 == 0:
                    if cell.fill.start_color.rgb in (None, "00000000"):
                        cell.fill = ALT_FILL
            row += 1
    widths = [8, 8, 10, 9, 8, 8, 10, 11, 10, 11, 10, 10, 10, 8, 8, 32, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_period_detail(wb, period, data):
    if not data:
        return
    ws = wb.create_sheet(f"{period}_detail")
    headers = ["Coin", "Strategy", "Train P", "Train R", "Train F1", "Train Fires",
               "Test P", "Test R", "Weight"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    row = 2
    for coin, d in data.items():
        if coin in DROPPED:
            continue
        weights = d.get("weights", {})
        for sk, s in d["strategyStats"].items():
            ws.cell(row=row, column=1, value=coin)
            ws.cell(row=row, column=2, value=sk)
            ws.cell(row=row, column=3, value=s["train"]["precision"]).number_format = "0.00"
            ws.cell(row=row, column=4, value=s["train"]["recall"]).number_format = "0.00"
            ws.cell(row=row, column=5, value=s["train"]["f1"]).number_format = "0.00"
            ws.cell(row=row, column=6, value=s["train"]["fires"])
            ws.cell(row=row, column=7, value=s["test"]["precision"]).number_format = "0.00"
            ws.cell(row=row, column=8, value=s["test"]["recall"]).number_format = "0.00"
            w = weights.get(sk, 0)
            wcell = ws.cell(row=row, column=9, value=w)
            wcell.number_format = "0.00"
            if w > 0:
                wcell.fill = GOOD_FILL
                wcell.font = ARIAL_BOLD
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                if cell.font.name != "Arial" or cell.font.bold is None:
                    cell.font = ARIAL if c != 9 or w == 0 else ARIAL_BOLD
                cell.border = BORDER
                if c >= 3:
                    cell.alignment = CENTER
            row += 1
    widths = [8, 10, 10, 10, 10, 11, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_readme(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("코인별 합성 전략 리포트", ARIAL_HEAD),
        ("", None),
        ("생성 방법:", ARIAL_BOLD),
        ("1) node scripts/composite_strategy_lab.mjs   # HTX kline 가져와서 학습 (8개 코인 × 3 봉)", ARIAL),
        ("2) node scripts/build_composite_presets.mjs  # web/lib/trading/composite_presets.ts 생성", ARIAL),
        ("3) python scripts/build_composite_report.py  # 이 파일 생성", ARIAL),
        ("", None),
        ("학습 방법론:", ARIAL_BOLD),
        ("- 5% 이상 하락 후 +20% 반등 → 그 저점을 '스윙 저점'으로 라벨링", ARIAL),
        ("- 8개 베이스 전략(ma, rsi, rsif, bb, bbf, donchian, vbf, zscore, momvol) 신호를 ±3봉 윈도우에서 정밀도 측정", ARIAL),
        ("- Train 80% / Test 20% walk-forward", ARIAL),
        ("- Weight = train precision (정규화)", ARIAL),
        ("- Threshold = 가중치 합의 50% 이상이면 매수", ARIAL),
        ("- 청산: 진입 후 최고가 대비 -10% 트레일링 스톱 또는 N봉 시간 청산", ARIAL),
        ("", None),
        ("HTX kline API 한계:", ARIAL_BOLD),
        ("- size=2000 단일 페이지만 지원 → 5년치는 1day만 가능, 4hour ≈ 333일, 60min ≈ 83일", ARIAL),
        ("- 시간봉 5년 데이터가 필요하면 Binance 페이지네이션 fetcher 추가 필요", ARIAL),
        ("", None),
        ("색상:", ARIAL_BOLD),
        ("- 초록: Test Sharpe ≥ 0.5 (out-of-sample 검증 통과)", ARIAL),
        ("- 빨강: Test Sharpe < 0 (검증 실패, 사용 비추)", ARIAL),
        ("", None),
        ("제외 코인: BCH, ZEC (수익성 부진으로 학습 대상 제외)", ARIAL_BOLD),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        if i == 1:
            cell.fill = HEAD_FILL
    ws.column_dimensions["A"].width = 110


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    all_data = {p: load_period(p) for p in PERIODS}
    if not any(all_data.values()):
        print("No composite_result_*.json found. Run lab script first.")
        return

    wb = Workbook()
    write_summary(wb.active, all_data)
    for p in PERIODS:
        write_period_detail(wb, p, all_data.get(p))
    write_readme(wb)

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
