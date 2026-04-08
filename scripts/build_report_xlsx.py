import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('data/coin_lab/presets.json', 'r') as f:
    presets = json.load(f)

wb = Workbook()

bold = Font(name='Arial', bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', start_color='1F4E78')
center = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
arial = Font(name='Arial')
hi_fill = PatternFill('solid', start_color='FFF2CC')

# ── Summary sheet ───────────────────────────────────────
ws = wb.active
ws.title = 'Summary'
headers = ['Coin', 'Symbol', 'Best Strategy', 'Params', 'Sharpe', 'CAGR', 'MDD', 'Calmar', 'Trades', 'Win%', 'Total Ret', 'Ann Vol', 'Trend R²', 'Hurst', 'BuyHold Ret']
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(1, c)
    cell.font = bold
    cell.fill = hdr_fill
    cell.alignment = center

for coin, p in presets.items():
    m, r = p['metrics'], p['regime']
    ws.append([
        coin, p['symbol'], p['strategy'], json.dumps(p['params']),
        m['sharpe'], m['cagr'], m['mdd'], m['calmar'], m['trades'], m['winRate'], m['totalReturn'],
        r['annualVol'], r['trendR2'], r['hurst'], r['buyHoldReturn'],
    ])

last = ws.max_row
for row in ws.iter_rows(min_row=2, max_row=last, max_col=len(headers)):
    for cell in row:
        cell.font = arial
        cell.border = border
for col, w in zip('ABCDEFGHIJKLMNO', [6, 12, 14, 32, 8, 8, 8, 8, 8, 8, 10, 9, 10, 8, 11]):
    ws.column_dimensions[col].width = w
for r in range(2, last + 1):
    for c in [6, 7, 10, 11, 12, 13, 15]:  # CAGR, MDD, Win%, Total, Vol, R², BH
        ws.cell(r, c).number_format = '0.0%'
    ws.cell(r, 5).number_format = '0.00'   # Sharpe
    ws.cell(r, 8).number_format = '0.00'   # Calmar
    ws.cell(r, 14).number_format = '0.00'  # Hurst

# Averages row
ws.cell(last + 2, 1, 'Avg').font = bold
for c, col in [(5, 'E'), (6, 'F'), (7, 'G'), (8, 'H'), (10, 'J'), (12, 'L')]:
    ws.cell(last + 2, c, f'=AVERAGE({col}2:{col}{last})')
    ws.cell(last + 2, c).font = bold
ws.cell(last + 2, 5).number_format = '0.00'
ws.cell(last + 2, 6).number_format = '0.0%'
ws.cell(last + 2, 7).number_format = '0.0%'
ws.cell(last + 2, 8).number_format = '0.00'
ws.cell(last + 2, 10).number_format = '0.0%'
ws.cell(last + 2, 12).number_format = '0.0%'

ws.freeze_panes = 'A2'

# ── Per-coin Top3 sheet ─────────────────────────────────
ws2 = wb.create_sheet('Top3 by Coin')
ws2.append(['Coin', 'Rank', 'Strategy', 'Params', 'Sharpe', 'CAGR', 'MDD', 'Trades', 'Win%'])
for c in range(1, 10):
    cell = ws2.cell(1, c)
    cell.font = bold
    cell.fill = hdr_fill
    cell.alignment = center

row = 2
for coin, p in presets.items():
    for i, t in enumerate(p['top3']):
        ws2.cell(row, 1, coin if i == 0 else '')
        ws2.cell(row, 2, i + 1)
        ws2.cell(row, 3, t['strategy'])
        ws2.cell(row, 4, json.dumps(t['params']))
        ws2.cell(row, 5, t['sharpe'])
        ws2.cell(row, 6, t['cagr'])
        ws2.cell(row, 7, t['mdd'])
        ws2.cell(row, 8, t['trades'])
        ws2.cell(row, 9, t['winRate'])
        if i == 0:
            for c in range(1, 10):
                ws2.cell(row, c).fill = hi_fill
        row += 1

for r in range(2, row):
    for c in range(1, 10):
        ws2.cell(r, c).font = arial
        ws2.cell(r, c).border = border
    ws2.cell(r, 5).number_format = '0.00'
    ws2.cell(r, 6).number_format = '0.0%'
    ws2.cell(r, 7).number_format = '0.0%'
    ws2.cell(r, 9).number_format = '0.0%'

for col, w in zip('ABCDEFGHI', [6, 5, 14, 38, 8, 8, 8, 8, 8]):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A2'

# ── Regime metrics sheet ────────────────────────────────
ws3 = wb.create_sheet('Regime')
ws3.append(['Coin', 'Annual Vol', 'Trend R²', 'Hurst', 'Buy&Hold Ret', 'Buy&Hold MDD', 'Regime Tag'])
for c in range(1, 8):
    cell = ws3.cell(1, c)
    cell.font = bold
    cell.fill = hdr_fill
    cell.alignment = center

def tag(r):
    if r['trendR2'] > 0.5:
        return 'Strong Trend'
    if r['trendR2'] < 0.2:
        return 'Choppy / Range'
    return 'Mixed'

for coin, p in presets.items():
    r = p['regime']
    ws3.append([coin, r['annualVol'], r['trendR2'], r['hurst'], r['buyHoldReturn'], r['buyHoldMdd'], tag(r)])

for rr in range(2, ws3.max_row + 1):
    for c in range(1, 8):
        ws3.cell(rr, c).font = arial
        ws3.cell(rr, c).border = border
    ws3.cell(rr, 2).number_format = '0.0%'
    ws3.cell(rr, 3).number_format = '0.00'
    ws3.cell(rr, 4).number_format = '0.00'
    ws3.cell(rr, 5).number_format = '0.0%'
    ws3.cell(rr, 6).number_format = '0.0%'
for col, w in zip('ABCDEFG', [6, 11, 10, 8, 12, 12, 16]):
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = 'A2'

out = '/sessions/friendly-awesome-bardeen/mnt/auto_coin/coin_strategy_report.xlsx'
wb.save(out)
print('saved', out)
